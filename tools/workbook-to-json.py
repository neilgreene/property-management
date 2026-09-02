#!/usr/bin/env python3
"""
Reads SDI analysis workbooks and emits the intake JSON the loader takes.

    python3 tools/workbook-to-json.py *.xlsm > batch.json
    node worker/tools/load-intake.js batch.json

WHY TWO STEPS. Reading .xlsm needs a spreadsheet library, and the worker
image has no dependency beyond the Postgres driver. Rather than drag one
in, the conversion happens wherever the file already is -- a laptop, the
host -- and what crosses into the database is plain JSON that a person can
open and read. The middle format is the point: when a released listing
says something surprising, you can look at exactly what was extracted
before anything touched it.

WHICH SHEET. The `Import` sheet, read by label. It is the only sheet that
carries the address, and its labels are identical across the workbooks
checked. `One Row` is a flatter export but has no address, so it cannot
stand alone.

WHAT IS DELIBERATELY NOT MAPPED TO A FIELD. "Schools Rating" and the
FAVORABLE/INSUFFICIENT deal score. Both are registered fair-housing
proxies (gov.prohibited_dimension): school ratings track catchment
demographics, and a composite score launders whatever went into it.
Offering either as something a buyer can filter or rank on is steering,
and intent is not required. They stay in `raw` -- the file is not edited,
and staff underwriting may legitimately consider schools -- and they never
become a column.
"""
import json
import re
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

MAPPING_VERSION = "v1"

# Address strings look like "401 NW 71st St, Kansas City, MO 64118".
ADDR = re.compile(r"^\s*(?P<street>.+?),\s*(?P<city>[^,]+?),\s*(?P<state>[A-Z]{2})\s+(?P<zip>\d{5})")


def read_import_sheet(ws):
    """Label -> value(s). Later columns are kept so 'Monthly | Yearly' pairs survive."""
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label:
            continue
        vals = [c for c in row[1:] if c is not None]
        if label not in out:
            out[label] = vals
    return out


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s or s.startswith("#"):        # #VALUE!, #REF! and friends
        return None
    try:
        return float(s)
    except ValueError:
        return None


def first(d, label, idx=0):
    v = d.get(label)
    if not v or len(v) <= idx:
        return None
    return v[idx]


def fnum(d, label, idx=0):
    return num(first(d, label, idx))


def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Import" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"{path}: no 'Import' sheet; is this an SDI analysis workbook?")
    d = read_import_sheet(wb["Import"])
    wb.close()

    raw = {}
    for k, v in d.items():
        raw[k] = [
            (x.isoformat() if isinstance(x, (datetime, date))
             else x if isinstance(x, (int, float, str)) else str(x))
            for x in v
        ]

    addr_s = str(first(d, "Address") or "")
    m = ADDR.match(addr_s)
    street = city = state = zipc = None
    if m:
        street, city, state, zipc = (m.group("street").strip(), m.group("city").strip(),
                                     m.group("state"), m.group("zip"))

    # Rent: the workbook gives a lower and an upper estimate, and the
    # 'Web' sheet -- what SDI itself shows publicly -- uses the midpoint.
    # Matching their own presentation rather than inventing a third
    # convention; both ends stay in raw.
    rent_hi = fnum(d, "Rent (upper)*")
    rent_lo = fnum(d, "Rent (lower)*")
    rent_m = None
    if rent_hi is not None and rent_lo is not None:
        rent_m = round((rent_hi + rent_lo) / 2, 2)
    elif rent_lo is not None:
        rent_m = rent_lo

    # Expenses. "Total Fixed Expenses" is the workbook's own operating
    # total and excludes the mortgage, which is what our opex_annual
    # means. HOA is carried separately here, so it is subtracted out.
    total_fixed_y = fnum(d, "Total Fixed Expenses", 1)
    hoa_y = fnum(d, "HOA or Fixed Costs", 1) or 0.0
    opex_y = None
    if total_fixed_y is not None:
        opex_y = round(total_fixed_y - hoa_y, 2)

    pm_pct = fnum(d, "Property Management Monthly (%)")
    vac = fnum(d, "Vacancy Rate")

    return {
        "source_file": path.split("/")[-1],
        "raw": raw,

        "street_address": street,
        "city": city,
        "state": state,
        "zip": zipc,
        # The workbook has no type field. Everything seen so far is a
        # single-family house, and "# of doors" tells us when it is not.
        "property_type": ("Single Family" if (fnum(d, "# of doors") or 1) <= 1
                          else "Multi-Family"),
        "beds": fnum(d, "Bedrooms"),
        "baths": fnum(d, "Bathrooms"),
        "sqft": fnum(d, "Square Feet"),
        "year_built": fnum(d, "Year Built"),
        "lot_sqft": fnum(d, "Lot size (sq ft)"),
        "garage_spaces": fnum(d, "Garage Size"),

        # The price an investor is being shown, not the seller's asking.
        # Asking, the suggested range and the all-in total stay in raw and
        # are summarised into internal notes below.
        "list_price": fnum(d, "Offer used for analysis"),

        "market_rent_monthly": rent_m,
        "gross_rent_annual": round(rent_m * 12, 2) if rent_m else None,
        "opex_annual": opex_y,
        "hoa_annual": hoa_y,
        "property_tax_annual": fnum(d, "Property Taxes", 1),
        "insurance_annual": fnum(d, "Insurance", 1),
        "maintenance_annual": fnum(d, "Repairs", 1),
        "management_fee_bps": int(round(pm_pct * 10000)) if pm_pct is not None else None,
        "vacancy_allowance_bps": int(round(vac * 10000)) if vac is not None else None,

        "description": (str(first(d, "General Notes About This Property") or "").strip()
                        or None),
        "internal_notes": "; ".join(x for x in [
            f"asking {fnum(d, 'Asking'):,.0f}" if fnum(d, "Asking") else None,
            (f"suggested {fnum(d, 'Suggested offer (low)'):,.0f}"
             f"-{fnum(d, 'Suggested offer (high)'):,.0f}"
             if fnum(d, "Suggested offer (low)") else None),
            f"all-in {fnum(d, 'Total Cost (estimated)'):,.0f}"
            if fnum(d, "Total Cost (estimated)") else None,
            f"metro {first(d, 'Metro')}" if first(d, "Metro") else None,
            f"DOM {fnum(d, 'DOM (days on market - TODAY)'):,.0f}"
            if fnum(d, "DOM (days on market - TODAY)") is not None else None,
        ] if x) or None,
    }


def main(argv):
    if len(argv) < 2:
        sys.exit(__doc__.strip())
    rows = [convert(p) for p in argv[1:]]
    print(json.dumps({
        "mapping_version": MAPPING_VERSION,
        "source_kind": "sdi_workbook",
        "source_file": (argv[1].split("/")[-1] if len(argv) == 2
                        else f"{len(rows)} workbooks"),
        "rows": rows,
    }, indent=2))


if __name__ == "__main__":
    main(sys.argv)
